#!/usr/bin/env python3
"""Check if Excel files contain age, height, and weight data"""

import openpyxl
from pathlib import Path
import sys

def check_excel_file(filepath):
    """Check a single Excel file for demographic info"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

        # Look for Subject Info sheet or similar
        sheet_names = wb.sheetnames

        # Check first sheet (usually has subject info)
        if sheet_names:
            sheet = wb[sheet_names[0]]

            # Search for Age, Height, Weight in first 50 rows
            age = None
            height = None
            weight = None
            dob = None

            for row in range(1, min(51, sheet.max_row + 1)):
                for col in range(1, min(10, sheet.max_column + 1)):
                    cell = sheet.cell(row, col)
                    cell_value = str(cell.value).lower() if cell.value else ""

                    # Check next cell for value
                    next_cell = sheet.cell(row, col + 1)
                    next_value = next_cell.value

                    if "age" in cell_value and "year" not in cell_value:
                        age = next_value
                    elif "height" in cell_value or "stature" in cell_value:
                        height = next_value
                    elif "weight" in cell_value and "body" in cell_value:
                        weight = next_value
                    elif "date of birth" in cell_value or "dob" in cell_value:
                        dob = next_value

            return {
                "filename": Path(filepath).name,
                "age": age,
                "height": height,
                "weight": weight,
                "dob": dob,
                "sheets": sheet_names
            }

        return None

    except Exception as e:
        print(f"Error reading {filepath}: {e}", file=sys.stderr)
        return None


if __name__ == "__main__":
    import sys

    cpet_data_dir = Path("../CPET_data")

    if len(sys.argv) > 1:
        # Check specific files
        files = [Path(f) for f in sys.argv[1:]]
    else:
        # Check a few sample files
        files = list(cpet_data_dir.glob("*.xlsx"))[:10]

    print("=" * 80)
    print("Excel 파일 인구통계 정보 확인")
    print("=" * 80)

    for filepath in files:
        if not filepath.exists():
            continue

        print(f"\n📄 {filepath.name}")
        result = check_excel_file(filepath)

        if result:
            print(f"   Age: {result['age']}")
            print(f"   Height: {result['height']}")
            print(f"   Weight: {result['weight']}")
            print(f"   DOB: {result['dob']}")
            print(f"   Sheets: {', '.join(result['sheets'][:3])}")
        else:
            print("   ❌ Failed to read file")
